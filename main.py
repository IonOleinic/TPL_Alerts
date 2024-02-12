import requests
from bs4 import BeautifulSoup
import time
import pytz
import os
import openpyxl
from statie import Statie
from alerta import Alerta
from stoc import Stoc
from threading import Thread
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import PyPDF2
from whatsapp import send_whatsapp_msg
import file_logger
import network
import processes

default_browser_exe = 'msedge.exe'
downloads_path = str(Path.home() / "Downloads")
whatsapp_group_id = 'HR4sPdnEGnI1vNGGehBmr4'



def check_ips():
    path = "IP echipamente statii.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # max_col = sheet_obj.max_column
    # max_row = sheet_obj.max_row
    max_row = 58
    stations_list = []
    for i in range(2, max_row+1):
        cell_nume_statie = sheet_obj.cell(i, 2).value
        cell_panou = sheet_obj.cell(i, 3).value
        cell_tvm = sheet_obj.cell(i, 4).value
        cell_camera = sheet_obj.cell(i, 5).value
        cell_switch = sheet_obj.cell(i, 6).value
        if (cell_nume_statie != None):
            station = Statie(cell_nume_statie, cell_panou,
                             cell_tvm, cell_camera, cell_switch)
            stations_list.append(station)
    for i in range(len(stations_list)):
        stations_list[i].check_station()


def check_alerts():
    skayo_session = requests.session()
    alert_list_already_send = []
    refresh_alerte = 25
    refresh_alerte_default = 25
    wait_time = 30*60

    def login_Skayo():
        try:
            if(network.check_local_network_conn()==False):
                file_logger.log(f"(Login Skayo) Se asteapta conexiunea la retea...")
                while(network.check_local_network_conn()==False):
                    time.sleep(60)
                file_logger.log(f"(Login Skayo) conexiunea la retea restabilita cu succes.")
            res = skayo_session.get(
                "http://192.168.95.93/Skayo_CFM/Authentication.aspx?ReturnUrl=/Skayo_CFM/Default.aspx")
            soup = BeautifulSoup(res.content, 'html.parser')
            ctl03_ctl00_TSM = soup.find(
                'input', id="ctl03_ctl00_TSM").attrs["value"]
            __VIEWSTATE = soup.find('input', id="__VIEWSTATE").attrs["value"]
            __VIEWSTATEGENERATOR = soup.find(
                'input', id="__VIEWSTATEGENERATOR").attrs["value"]
            __VIEWSTATEENCRYPTED = soup.find(
                'input', id="__VIEWSTATEENCRYPTED").attrs["value"]
            __EVENTVALIDATION = soup.find(
                'input', id="__EVENTVALIDATION").attrs["value"]
            ctl03_cboCompany_ClientState = soup.find(
                'input', id="__EVENTVALIDATION").attrs["value"]
            txt_username = "admin"
            txt_password = "ticketing"
            tib_authentication = "Autentificare"
            payload = {
                "ctl03_ctl00_TSM": ctl03_ctl00_TSM,
                "__VIEWSTATE": __VIEWSTATE,
                "__VIEWSTATEGENERATOR": __VIEWSTATEGENERATOR,
                "__VIEWSTATEENCRYPTED": __VIEWSTATEENCRYPTED,
                "__EVENTVALIDATION": __EVENTVALIDATION,
                "ctl03_cboCompany_ClientState": ctl03_cboCompany_ClientState,
                "ctl03$txtUserName": txt_username,
                "ctl03$txtPassword": txt_password,
                "ctl03$tibAuthentication": tib_authentication
            }
            res = skayo_session.post(
                "http://192.168.95.93/Skayo_CFM/Authentication.aspx?ReturnUrl=/Skayo_CFM/Default.aspx", data=payload)
            return True
        except Exception as e:
            file_logger.log(e)
            return False
 
    def check_if_alert_in(alert_id, alert_list):
        for i in range(len(alert_list)):
            if (alert_id == alert_list[i].id):
                return True
        return False

    def delete_expired_alerts(alert_list):
        result_list = []
        for i in range(len(alert_list)):
            if (alert_list[i].ttl > 0):
                result_list.append(alert_list[i])
            else:
                file_logger.log(
                    f'A expirat timpul pentru alerta {alert_list[i].nume} {alert_list[i].data}')
        return result_list

    login_Skayo()
    
    while True:
        try:
            now = datetime.now()
            if (now.hour >= 6 and now.hour <= 18):
                refresh_alerte = 10
            else:
                refresh_alerte = refresh_alerte_default
            default_alert_ttl = wait_time/refresh_alerte
            response = skayo_session.get(
                'http://192.168.95.93/Skayo_CFM/AVM/Manage/AlertListView.aspx')
            soup = BeautifulSoup(response.content, 'html.parser')
            table_id = 'ctl00_cphContent_gridAlerts_ctl00'
            table_tag = soup.find('table', id=table_id)
            if (table_tag):
                alert_finded = False
                for i in range(10):
                    row_tag = table_tag.find('tr', id=(table_id+"__"+str(i)))
                    td_list = row_tag.find_all('td', class_='col-lg-1')
                    tip_alerta = row_tag.find('td', class_='col-lg-3').text
                    tip_eroare = row_tag.find('td', class_='rowspan col-lg-7').text
                    nume_TVM = td_list[0].text
                    data_alerta = td_list[1].text
                    alert_id = nume_TVM+data_alerta[0:10]
                    if ('Defect hardware' in tip_alerta):
                        alert_finded = True
                        if("BNR" in tip_eroare.upper()):
                            tip_eroare="Eroare BNR"
                        elif("POS" in tip_eroare.upper()):
                            tip_eroare="Eroare POS"
                        elif("CARDDISPENSER" in tip_eroare.upper()):
                            tip_eroare="Eroare imprimanta carduri"
                        elif("RECEIPTPRINTER" in tip_eroare.upper()):
                            tip_eroare="Eroare imprimanta chitante"
                        elif("QR" in tip_eroare.upper()):
                            tip_eroare="Eroare imprimanta bilete"
                        else:
                            tip_eroare="Eroare generala"
                        print(
                            f'\nDefect Hardware detectat!!!\n{nume_TVM}\n{data_alerta}\n{tip_eroare}\n{tip_alerta}')
                        if (check_if_alert_in(alert_id, alert_list_already_send) == False):
                            if (now.hour < 5 or now.hour > 21):
                                alert_ttl = 60*60/refresh_alerte
                            else:
                                alert_ttl = default_alert_ttl
                            new_alert = Alerta(
                                alert_id, nume_TVM, data_alerta, tip_alerta, alert_ttl)
                            file_logger.log(
                                f"\nTPL Suceava Skayo TVM Alert\n{nume_TVM}\n{data_alerta}{tip_alerta}{tip_eroare}")
                            send_whatsapp_msg(
                                "Echipa racheta", f"TPL Suceava Skayo TVM Alert\n{nume_TVM}\n{data_alerta}{tip_alerta}{tip_eroare}", "alerta")
                            alert_list_already_send.append(new_alert)
                        else:
                            print('Mesaj deja trimis pe Whatsap')
                if (alert_finded == False):
                    print("Nimic gasit.")
                alert_list_already_send = delete_expired_alerts(
                    alert_list_already_send)
            else:
                file_logger.log("Eroare logare Skayo. Se incearca din nou...")
                login_Skayo()

            for i in range(len(alert_list_already_send)):
                alert_list_already_send[i].ttl -= 1

            print(f"Refresh dupa {refresh_alerte} sec...")
            print('-------------------------------------')
            time.sleep(refresh_alerte)
        except Exception as e:
            file_logger.log(e)
            login_Skayo()


def check_stocks():

    def list_to_str(list, delimiter):
        result = ''
        for i in range(len(list)):
            result += list[i]+delimiter
        return result

    def parse_pdf(pdf_name):
        if (pdf_name == None):
            file_logger.log(
                f"[PARSE] File '{pdf_name}' not exist or was removed.")
            return
        else:
            try:
                pdf_path = downloads_path+'/'+pdf_name
                if os.path.exists(pdf_path):
                    with open(pdf_path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        num_pages = len(pdf_reader.pages)
                        list_to_collect = []
                        for i in range(num_pages):
                            page = pdf_reader.pages[i]
                            text = page.extract_text()
                            lines = text.splitlines()
                            for i in range(10, len(lines)):
                                line = lines[i].split(' ')
                                if (len(line) > 30):
                                    nume_TVM = list_to_str(line[30:], ' ')
                                    if ("Stefan" in nume_TVM):
                                        nume_TVM = "Colegiul Stefan cel Mare"
                                    nr_bancnote = int(line[11])
                                    if (nr_bancnote > 399):
                                        stoc = Stoc(nume_TVM, nr_bancnote)
                                        list_to_collect.append(stoc)

                        if (len(list_to_collect) == 0):
                            file_logger.log("Nu sunt TVM-uri de colectat.")
                        else:
                            whatsapp_text = ""
                            for stoc in list_to_collect:
                                whatsapp_text += stoc.nume_TVM + \
                                    " "+str(stoc.nr_bancnote)+"\n"
                            file_logger.log(
                                f"\nTPL Suceava Stocuri Bancnote:\n{whatsapp_text}")
                            send_whatsapp_msg(
                                "Echipa racheta", f"TPL Suceava Stocuri Bancnote:\n{whatsapp_text}", "situatie stocuri")
                else:
                    file_logger.log("PDF File Not Found.")
            except Exception as e:
                file_logger.log(e)

    def download_pdf_stocuri():
        def download_pdf():
            try:
                opts = webdriver.FirefoxOptions()
                opts.add_argument("--headless")
                driver = webdriver.Firefox(options=opts)
                time.sleep(1)
                driver.get(
                    "http://192.168.95.93/Skayo_CFM/Authentication.aspx?ReturnUrl=/Skayo_CFM/Default.aspx")
                username_field = driver.find_element(
                    'name', 'ctl03$txtUserName')
                username_field.send_keys('admin')
                time.sleep(1)
                password_field = driver.find_element(
                    'name', 'ctl03$txtPassword')
                password_field.send_keys('ticketing')
                time.sleep(1)
                login_form = driver.find_element(
                    'name', 'ctl03$tibAuthentication')
                login_form.click()
                time.sleep(1)
                driver.get(
                    "http://192.168.95.93/Skayo_CFM/Reporting/Local/ReportCreator.aspx?RID=428")
                generate_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.ID, "ctl00_cphContent_tibGenerate"))
                )
                generate_btn.click()
                time.sleep(10)
                driver.switch_to.frame("Situatie stocuri curenta")
                btn_export_pdf = WebDriverWait(driver, 35).until(
                    EC.element_to_be_clickable((By.ID, "btnExportPdf"))
                )
                btn_export_pdf.click()
                time.sleep(25)  # a suficient time for downloading pdf
                driver.quit()
                return True
            except Exception as e:
                file_logger.log(e)
                return False
        while (download_pdf() == False):
            file_logger.log(
                f"Eroare descarcare pdf stocuri.Se incearca din nou dupa...5 sec")
            # try to download after 5 sec
            processes.pkill("firefox.exe")
            time.sleep(5)

    def find_pdf(partial_name):
        full_list = os.listdir(downloads_path)
        latest_date = 0
        finded_file = None
        for file in full_list:
            if (partial_name in file):
                file_path = downloads_path+'/'+file
                file_name, file_extension = os.path.splitext(file_path)
                if (file_extension == '.pdf'):
                    if (os.path.getctime(file_path) > latest_date):
                        latest_date = os.path.getctime(file_path)
                        finded_file = file
        return finded_file

    def delete_pdf(pdf_name):
        if (pdf_name == None):
            file_logger.log(
                f"[DELETE] File '{pdf_name}' not exist or was removed.")
            return
        pdf_path = downloads_path+'/'+pdf_name
        if os.path.exists(pdf_path):
            try:
                os.remove(pdf_path)
                file_logger.log(f"FILE '{pdf_name}' was deleted")
            except OSError as e:
                file_logger.log(e)
        else:
            file_logger.log(f"[DELETE] The file '{pdf_name}' does not exist")

    refresh_stocuri = 3600
    while (True):
        try:
            if(network.check_local_network_conn()==False):
                file_logger.log(f"(Verificare Stocuri) Se asteapta conexiunea la retea...")
                while(network.check_local_network_conn()==False):
                    time.sleep(60)
                file_logger.log(f"(Verificare Stocuri) conexiunea la retea restabilita cu succes.")
            now = datetime.now()
            if (now.hour >= 6 and now.hour <= 20):
                download_pdf_stocuri()
                time.sleep(1)
                pdf_name = find_pdf('Situatie stocuri curenta')
                time.sleep(1)
                parse_pdf(pdf_name)
                delete_pdf(pdf_name)
                print()
        except Exception as e:
            file_logger.log(e)
        time.sleep(refresh_stocuri)


def check_mobile_messages():
    tpl_mobile_session = requests.session()
    def login_tpl_mobile():
        try:
            if (network.check_internet_conn() == False):
                print(f"(Login TPL Mobile) Se asteapta conexiunea la retea...")
                while (network.check_internet_conn() == False):
                    time.sleep(60)
                print(f"(Login TPL Mobile) conexiunea la retea restabilita cu succes.")
            payload = {"username": "ion.oleinic21@gmail.com",
                    "password": "Mazzeratti123"}
            jwt_token = None
            response = tpl_mobile_session.post(
                "https://mobile.tplsv.ro:9090/api/authenticate", json=payload)
            json_data = response.json()
            jwt_token = json_data.get('id_token', None)
            return jwt_token
        except Exception as e:
            print(e)
            return None
    
    def get_messages_tpl_mobile(jwt_token):
        headers = {
            'Authorization': f'Bearer {jwt_token}'
        }
        response = tpl_mobile_session.get(
            "https://mobile.tplsv.ro:9090/api/user-message-threads?page=0&size=20&type=0&email=&subject=&sort=createdDate,desc&sort=id", headers=headers)
        response.raise_for_status()
        messages = response.json()
        return messages

    jwt_token=login_tpl_mobile()
    refresh_mobile_msgs=1200
    messages_already_sent=[]
    while(True):
        try:
            now = datetime.now()
            if (now.hour >= 6 and now.hour <= 20):
                messages=get_messages_tpl_mobile(jwt_token)
                for message in messages:
                    if ((message['isSeen'] == False) and (message['id'] not in messages_already_sent)):
                        # message should be send on whatsapp
                        userEmail = message['userEmail']
                        datetime_object_utc = datetime.strptime( message['createdDate'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=pytz.UTC)
                        datetime_object_adjusted = datetime_object_utc.astimezone(pytz.timezone('Europe/Bucharest'))
                        createdDate = datetime_object_adjusted.strftime("%d.%m.%Y %H:%M")
                        subject = message['subject']
                        message_text = message['userMessages'][0]['text'].replace(
                            '\n', ' ')
                        message_to_send = f'TPL Mobile Message\nEmail: {userEmail}\nData: {createdDate}\nSubiect: {subject}\nMesaj: {message_text}'
                        file_logger.log(f'\n{message_to_send}')
                        send_whatsapp_msg(["+373 671 06 737","Simona"],message_to_send,'reclamatie mobile')
                        messages_already_sent.append(message['id'])
            time.sleep(refresh_mobile_msgs)

        except requests.exceptions.HTTPError as err:
            if err.response.status_code == 401:
                file_logger.log('Eroare logare TPL Mobile. Se incearca din nou...') 
                jwt_token=login_tpl_mobile()
            else:
                file_logger.log(err) 
        except Exception as err:
            file_logger.log("TPL Mobile Thread CRASHED.")
            file_logger.log(err)
            send_whatsapp_msg("+373 671 06 737","TPL Mobile Thread CRASHED.",'reclamatie mobile')
            raise Exception("TPL Mobile Thread CRASHED.") #STOP
            
        
def main():
    file_logger.init_logs()
    # send_whatsapp_msg("+373 671 06 737", "Start Script Alerte", 'start script')
    Thread(target=check_stocks).start()
    Thread(target=check_ips).start()
    Thread(target=check_alerts).start()
    Thread(target=check_mobile_messages()).start()


main()
